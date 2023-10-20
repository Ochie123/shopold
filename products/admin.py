from django.contrib import admin

# Register your models here.
# /products/admin.py
from copy import copy
from openpyxl import Workbook
from openpyxl.styles import Alignment, NamedStyle, builtins
from openpyxl.styles.numbers import FORMAT_NUMBER
from openpyxl.writer.excel import save_virtual_workbook

from django.contrib import admin
from django import forms
from django.db import models
from django.http.response import HttpResponse
from django.template.loader import render_to_string
from django.utils.html import mark_safe
from django.utils.translation import gettext_lazy as _
from ordered_model.admin import OrderedTabularInline, OrderedInlineModelAdminMixin

from .models import Product, ProductImage, SearchTerm, ProductView, Tag
from categories.models import Category

admin.site.site_header = 'SVG Craft Administration'
admin.site.site_header_color = "purple" 
admin.site.module_caption_color = "grey"

admin.site.register(Tag)
class TagsAdmin(admin.ModelAdmin):
    fieldsets = ((_("Tag"), {"fields": ("title",)}),)
    prepopulated_fields = {"slug": ("title",)}

class ProductForm(forms.ModelForm):
    categories = forms.ModelMultipleChoiceField(
        label=_("Categories"), 
        queryset=Category.objects.all(), 
        widget=forms.CheckboxSelectMultiple(), 
        required=True,
    )
    
    class Meta:
        model = Product
        fields = "__all__"

    def __init__(self, *args, **kwargs): 
        super().__init__(*args, **kwargs)

      

def export_xlsx(modeladmin, request, queryset):
    wb = Workbook()
    ws = wb.active
    ws.title = "Products"

    number_alignment = Alignment(horizontal="right")
    wb.add_named_style(
        NamedStyle(
            "Identifier", alignment=number_alignment, number_format=FORMAT_NUMBER
        )
    )
    wb.add_named_style(
        NamedStyle("Normal Wrapped", alignment=Alignment(wrap_text=True))
    )

    class Config:
        def __init__(
            self,
            heading,
            width=None,
            heading_style="Headline 1",
            style="Normal Wrapped",
            number_format=None,
        ):
            self.heading = heading
            self.width = width
            self.heading_style = heading_style
            self.style = style
            self.number_format = number_format

    column_config = {
        "A": Config("ID", width=10, style="Identifier"),
        "B": Config("Title", width=30),
        "C": Config("Description", width=60),
        "D": Config("Price", width=15, style="Currency", number_format="#,##0.00 $"),
        "E": Config("Preview", width=100, style="Hyperlink"),
    }

    # Set up column widths, header values and styles
    for col, conf in column_config.items():
        ws.column_dimensions[col].width = conf.width

        column = ws[f"{col}1"]
        column.value = conf.heading
        column.style = conf.heading_style

    # Add products
    for obj in queryset.order_by("pk"):
        project_images = obj.productimage_set.all()[:1]
        url = ""
        if project_images:
            url = project_images[0].image.url

        data = [obj.pk, obj.title, obj.description, obj.price, url]
        ws.append(data)

        row = ws.max_row
        for row_cells in ws.iter_cols(min_row=row, max_row=row):
            for cell in row_cells:
                conf = column_config[cell.column_letter]
                cell.style = conf.style
                if conf.number_format:
                    cell.number_format = conf.number_format

    mimetype = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    charset = "utf-8"
    response = HttpResponse(
        content=save_virtual_workbook(wb),
        content_type=f"{mimetype}; charset={charset}",
        charset=charset,
    )
    response["Content-Disposition"] = "attachment; filename=products.xlsx"
    return response


export_xlsx.short_description = _("Export XLSX")


ZERO = "zero"
ONE = "one"
MANY = "many"


class ImageFilter(admin.SimpleListFilter):
    # Human-readable title which will be displayed in the
    # right admin sidebar just above the filter options.
    title = _("images")

    # Parameter for the filter that will be used in the
    # URL query.
    parameter_name = "images"

    def lookups(self, request, model_admin):
        """
        Returns a list of tuples, akin to the values given for
        model field choices. The first element in each tuple is the
        coded value for the option that will appear in the URL
        query. The second element is the human-readable name for
        the option that will appear in the right sidebar.
        """
        return (
            (ZERO, _("Has no images")),
            (ONE, _("Has one image")),
            (MANY, _("Has more than one image")),
        )

    def queryset(self, request, queryset):
        """
        Returns the filtered queryset based on the value
        provided in the query string and retrievable via
        `self.value()`.
        """
        qs = queryset.annotate(num_images=models.Count("productimage"))

        if self.value() == ZERO:
            qs = qs.filter(num_images=0)
        elif self.value() == ONE:
            qs = qs.filter(num_images=1)
        elif self.value() == MANY:
            qs = qs.filter(num_images__gte=2)
        return qs


class ProductImageInline(OrderedTabularInline):
    model = ProductImage
    extra = 0
    fields = ("get_image_preview", "image", "order", "move_up_down_links")
    readonly_fields = ("get_image_preview", "order", "move_up_down_links")
    ordering = ("order",)

    def get_image_preview(self, obj):
        photo_preview = render_to_string(
            "admin/products/includes/photo-preview.html",
            {"image": obj, "product": obj.product},
        )
        return mark_safe(photo_preview)

    get_image_preview.short_description = _("Preview")
@admin.register(Product)
class ProductAdmin(OrderedInlineModelAdminMixin, admin.ModelAdmin):
    list_display = ["first_image",'publish','status', "title", "has_description", "has_file", "price"]
    list_display_links = ["first_image", "title"]
    list_editable = ["price"]
    list_filter = [ImageFilter, 'bestseller', 'toprated', 'status', ]

    actions = [export_xlsx]

    fieldsets = ((_("Product"), {"fields": ("categories","tags",'toprated', 'bestseller',"title", "slug", "description", "price","file", "status")}),)
    prepopulated_fields = {"slug": ("title",)}
    inlines = [ProductImageInline]

    def first_image(self, obj):
        project_images = obj.productimage_set.all()[:1]
        if project_images.count() > 0:
            image_preview = render_to_string(
                "admin/products/includes/photo-preview.html",
                {"image": project_images[0], "product": obj},
            )
            return mark_safe(image_preview)
        return ""

    first_image.short_description = _("Preview")

    def has_description(self, obj):
        return bool(obj.description)

    has_description.short_description = _("Has description?")
    has_description.boolean = True

    def has_file(self, obj):
        return bool(obj.file)

    has_file.short_description = _("Has File?")
    has_file.boolean = True

@admin.register(SearchTerm)
class SearchTermAdmin(admin.ModelAdmin):
    list_display = ('__unicode__','ip_address','search_date', 'user') 
    list_filter = ('ip_address', 'q', 'user', 'search_date')
    exclude = ('user',)

@admin.register(ProductView)
class ProductViewAdmin(admin.ModelAdmin):
    list_display = ['product','ip_address', 'user', 'date', 'tracking_id']
    list_filter = ['product', 'user', 'date']
    exclude = ('user',)
  
